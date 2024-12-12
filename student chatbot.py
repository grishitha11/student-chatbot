import openpyxl
from openpyxl import Workbook
import tkinter as tk
from tkinter import messagebox
from tkinter import scrolledtext

def create_excel_file(file_name):
    # Create an Excel file to store student details if it doesn't already exist
    try:
        wb = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Student Details"
        sheet.append(["Name", "Roll Number", "Department", "Email", "Phone Number"])
        wb.save(file_name)
    except PermissionError:
        messagebox.showerror("Permission Error", f"Unable to access {file_name}. Close the file if it is open and try again.")
        exit()

def save_student_details(file_name, student_data):
    # Append student details to the Excel file
    try:
        wb = openpyxl.load_workbook(file_name)
        sheet = wb.active
        sheet.append(student_data)
        wb.save(file_name)
    except PermissionError:
        messagebox.showerror("Permission Error", f"Unable to access {file_name}. Close the file if it is open and try again.")

def chatbot_interface():
    def ask_question():
        question = chat_input.get()
        if question.strip() == "":
            return

        chat_display.config(state=tk.NORMAL)
        chat_display.insert(tk.END, f"You: {question}\n")
        chat_input.delete(0, tk.END)

        # Handle responses based on question
        global current_step
        global student_data

        if current_step == 0:
            student_data["Name"] = question
            chat_display.insert(tk.END, "Bot: What is your Roll Number?\n")
            current_step += 1
        elif current_step == 1:
            student_data["Roll Number"] = question
            chat_display.insert(tk.END, "Bot: What is your Department?\n")
            current_step += 1
        elif current_step == 2:
            student_data["Department"] = question
            chat_display.insert(tk.END, "Bot: What is your Email?\n")
            current_step += 1
        elif current_step == 3:
            student_data["Email"] = question
            chat_display.insert(tk.END, "Bot: What is your Phone Number?\n")
            current_step += 1
        elif current_step == 4:
            student_data["Phone Number"] = question
            save_student_details("Student_Details.xlsx", list(student_data.values()))
            chat_display.insert(tk.END, "Bot: Thank you! Your details have been saved.\n")
            current_step = 0
            student_data = {}

        chat_display.config(state=tk.DISABLED)

    create_excel_file("Student_Details.xlsx")

    # Initialize student data
    global student_data
    student_data = {}
    global current_step
    current_step = 0

    # GUI Setup
    root = tk.Tk()
    root.title("Student Details Chatbot")

    # Chat display area
    chat_display = scrolledtext.ScrolledText(root, wrap=tk.WORD, width=50, height=20, state=tk.DISABLED)
    chat_display.grid(row=0, column=0, columnspan=2, padx=10, pady=10)

    # Enable writing to the chat display
    chat_display.config(state=tk.NORMAL)
    chat_display.insert(tk.END, "Bot: Hello! What is your Name?\n")
    chat_display.config(state=tk.DISABLED)

    # Input area
    chat_input = tk.Entry(root, width=40)
    chat_input.grid(row=1, column=0, padx=10, pady=5)

    # Send button
    send_button = tk.Button(root, text="Send", command=ask_question)
    send_button.grid(row=1, column=1, padx=5, pady=5)

    # Run the GUI
    root.mainloop()

if __name__ == "__main__":
    chatbot_interface()
